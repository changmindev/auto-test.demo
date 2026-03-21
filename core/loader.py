from pathlib import Path
import openpyxl
from config import START_TC_ID, END_TC_ID
from utils.parsers import normalize_header, find_header_col


def load_cases(path: Path, sheet_name: str) -> list:
    """
    엑셀 파일에서 TC 목록을 읽어 START_TC_ID ~ END_TC_ID 범위로 필터링 후 반환.

    필수 컬럼: No. / 전제조건 / 테스트 항목 / 기대결과
    선택 컬럼: 자동화실행 / 자동화사유 / 자동화메모
    """
    if not path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"시트를 찾을 수 없습니다: '{sheet_name}' / 사용 가능 시트: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    headers = {}
    for c in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(1, c).value)
        if header:
            headers[header] = c

    print("사용 가능한 시트:", wb.sheetnames)
    print("실제 읽는 시트:", sheet_name)
    print("읽은 헤더:", list(headers.keys()))

    col_tc_id     = find_header_col(headers, "No.", "NO.", "No")
    col_state     = find_header_col(headers, "전제조건", "기능", "상태")
    col_test_item = find_header_col(headers, "테스트 항목", "테스트항목")
    col_expected  = find_header_col(headers, "기대결과", "기대 결과")
    col_auto_flag = find_header_col(headers, "자동화실행", "자동화 실행", "자동화")
    col_reason    = find_header_col(headers, "자동화사유", "자동화 사유",
                                    "비고\n(오류보고내역)", "비고(오류보고내역)")
    col_memo      = find_header_col(headers, "자동화메모", "자동화 메모",
                                    "검증결과\n(P,F,NT,NA)", "검증결과(P,F,NT,NA)")

    missing = [name for name, col in [
        ("No.", col_tc_id), ("전제조건", col_state),
        ("테스트 항목", col_test_item), ("기대결과", col_expected),
    ] if col is None]
    if missing:
        raise ValueError(f"필수 컬럼 없음: {missing} / 현재 헤더={list(headers.keys())}")

    rows = []
    for r in range(2, ws.max_row + 1):
        tc_id     = ws.cell(r, col_tc_id).value
        test_item = ws.cell(r, col_test_item).value
        expected  = ws.cell(r, col_expected).value

        if not tc_id or not test_item or not expected:
            continue

        rows.append({
            "tc_id":     str(tc_id).strip(),
            "state":     str(ws.cell(r, col_state).value or "").strip(),
            "test_item": str(test_item).strip(),
            "expected":  str(expected).strip(),
            "auto_flag": str(ws.cell(r, col_auto_flag).value if col_auto_flag else "AUTO").strip().upper(),
            "reason":    str(ws.cell(r, col_reason).value   if col_reason    else "").strip(),
            "memo":      str(ws.cell(r, col_memo).value     if col_memo      else "").strip(),
        })

    return _filter_by_range(rows, START_TC_ID, END_TC_ID)


def _filter_by_range(rows: list, start: str, end: str) -> list:
    """START_TC_ID ~ END_TC_ID 범위 슬라이싱"""
    if start:
        result, started = [], False
        for row in rows:
            if row["tc_id"] == start:
                started = True
            if started:
                result.append(row)
            if end and row["tc_id"] == end:
                break
        return result

    if end:
        result = []
        for row in rows:
            result.append(row)
            if row["tc_id"] == end:
                break
        return result

    return rows
